import time

import openpyxl
import pandas as pd

# get file
FILE = r'filepath'

# used for writing to the data
src_file = openpyxl.load_workbook(filename=FILE, read_only=False)
write_data = src_file['employee_data']
write_schedule = src_file['schedule']

# used for reading the data
read_data = pd.read_excel(FILE, 'employee_data')
read_schedule = pd.read_excel(FILE, 'schedule')


def check(col_end, name, skip_day, check_day, current_day):
    temp_col = 2
    temp_row = current_day + 2
    if not check_day:
        for x in range(col_end):
            if write_schedule.cell(row=temp_row, column=temp_col).value == name:
                skip_day = True
                # print(f'FOUND NAME:{skip_day}')
            else:
                temp_col += 1
        check_day = True
    # print(f'Was the day checked: {check_day}')
    return skip_day


#
#
#
for emp_row in range(len(read_data)):
    #
    #
    # assign variables
    EMP_NAME = read_data.iloc[emp_row, 0]  # employee's name
    days_list = read_data.iloc[emp_row, 2]  # employee's days available (Sunday - Saturday)
    shift_pref = read_data.iloc[emp_row, 3]  # employee's shift preference (morning \ night)
    station_pref = read_data.iloc[emp_row, 4]  # station preference ~ (T1: random, T2: set)

    print('\n'
          f'Name: {EMP_NAME}\n'
          f'Days: {days_list}\n'
          f'Shift: {shift_pref}\n')
    #
    #
    # using these timers to count variables
    SLOTS_COUNT = 5
    POSITION_TIMER = 5
    LOOP_COUNTER = 0
    #
    #
    # reference schedules
    while SLOTS_COUNT != 0:
        LOOP_COUNTER += 1
        for day in range(len(read_schedule)):
            CURRENT_DAY = read_schedule.iloc[day, 0]
            NEXT_ROW = False
            SKIP_DAY = False
            CHECK_DAY = False
            COL = 2
            COLUMN_END = 14

            print(f'Day:{day}')
            #
            #
            #
            # - CHECK  FUNCTION -
            SHOULD_SKIP_DAY = check(COLUMN_END, EMP_NAME, SKIP_DAY, CHECK_DAY, day)
            #
            #
            #
            # PRIORITY
            if LOOP_COUNTER < 2:
                #
                #
                #
                while not NEXT_ROW:
                    if SHOULD_SKIP_DAY:
                        print('SKIP DAY or COL END REACHED')
                        NEXT_ROW = True
                    #
                    #
                    #
                    elif COL > COLUMN_END:
                        print('COL END REACHED')
                        NEXT_ROW = True
                    #
                    #
                    #
                    elif write_schedule.cell(row=day + 2, column=COL).value == 0:
                        # MORNING preference employee handling
                        if CURRENT_DAY in days_list:
                            #
                            #
                            #
                            if shift_pref == 'morning':
                                if COL % 2 == 0:
                                    write_schedule.cell(row=day + 2, column=COL).value = EMP_NAME
                                    SLOTS_COUNT -= 1
                                    write_data.cell(row=emp_row + 2, column=7).value = SLOTS_COUNT
                                    src_file.save(filename=FILE)
                                    NEXT_ROW = True
                                else:
                                    COL += 1
                            #
                            #
                            #
                            # NIGHT preference employee handling
                            elif shift_pref == 'night':
                                if COL % 2 != 0:
                                    write_schedule.cell(row=day + 2, column=COL).value = EMP_NAME
                                    SLOTS_COUNT -= 1
                                    write_data.cell(row=emp_row + 2, column=7).value = SLOTS_COUNT
                                    src_file.save(filename=FILE)
                                    NEXT_ROW = True
                                else:
                                    COL += 1
                            #
                            #
                            #
                            else:
                                NEXT_ROW = True
                        #
                        #
                        # - if current day is not in days_list -
                        else:
                            NEXT_ROW = True
                    #
                    #
                    #
                    else:
                        COL += 1
                #
                #
                # FAILED PRIORITY
            else:
                print("FAILED PRIORITY")
                while not NEXT_ROW:
                    if SHOULD_SKIP_DAY:
                        print('SKIP DAY')
                        NEXT_ROW = True

                    elif COL > COLUMN_END:
                        print(f'BIG_COL END REACHED (COL: {COL})')
                        NEXT_ROW = True

                    elif write_schedule.cell(row=day + 2, column=COL).value == 0 and SLOTS_COUNT != 0:

                        print('BIG_CELL FOUND')
                        write_schedule.cell(row=day + 2, column=COL).value = EMP_NAME
                        SLOTS_COUNT -= 1
                        write_data.cell(row=emp_row + 2, column=7).value = SLOTS_COUNT
                        src_file.save(filename=FILE)
                        NEXT_ROW = True
                    else:
                        COL += 1

print('\nSaving...')
time.sleep(3)
src_file.save(filename=FILE)
print("SCHEDULE COMPLETE")
